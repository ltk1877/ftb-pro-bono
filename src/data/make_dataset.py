# -*- coding: utf-8 -*-
import re
from pathlib import Path

import logging
import pandas as pd
from openpyxl import load_workbook


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize spreadsheet column names for consistent downstream access."""
    out = df.copy()
    out.columns = [re.sub(r"\s+", " ", str(c).strip()) if c is not None else f"unnamed_{i}" for i, c in enumerate(out.columns)]
    return out


def read_excel_sheet_openpyxl(path: Path, sheet_name: str) -> pd.DataFrame:
    """Read a workbook sheet directly with openpyxl to avoid pandas/openpyxl version issues."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    # Drop trailing blank workbook columns so they do not become unusable dataframe fields.
    useful_cols = [i for i, value in enumerate(header) if value is not None]
    names = [header[i] for i in useful_cols]
    # Some sheets have a long tail of fully-blank rows past the real data (formatting artifacts);
    # skip any row that has no populated cells at all rather than assuming it matches the header width.
    data = [[row[i] for i in useful_cols] for row in rows if any(v is not None for v in row)]
    return clean_columns(pd.DataFrame(data, columns=names))


def read_sheet_from_workbooks(paths: list, sheet_name: str) -> pd.DataFrame:
    """Read the same-named sheet from multiple workbooks (e.g. successive MMG releases) and concatenate them."""
    frames = [read_excel_sheet_openpyxl(p, sheet_name) for p in paths]
    return pd.concat(frames, ignore_index=True)


def main(input_filepath, output_filepath):
    """ Runs data processing scripts to turn raw data from (../raw) into
        cleaned data ready to be analyzed (saved in ../processed).
    """
    logger = logging.getLogger(__name__)
    logger.info('making final data set from raw data')


if __name__ == '__main__':
    # click/dotenv are only needed for this CLI stub, not for the reusable functions above,
    # so they're imported here rather than at module level.
    import click
    from dotenv import find_dotenv, load_dotenv

    main = click.command()(
        click.argument('input_filepath', type=click.Path(exists=True))(
            click.argument('output_filepath', type=click.Path())(main)
        )
    )

    log_fmt = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    logging.basicConfig(level=logging.INFO, format=log_fmt)

    # not used in this stub but often useful for finding various files
    project_dir = Path(__file__).resolve().parents[2]

    # find .env automagically by walking up directories until it's found, then
    # load up the .env entries as environment variables
    load_dotenv(find_dotenv())

    main()
